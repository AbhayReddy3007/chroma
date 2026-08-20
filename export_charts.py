"""
export_charts.py — Export step9 all_final_charts JSON to Excel
===============================================================
Adds a "Filed By" column (from BigQuery patent_discovery table) beside
each Patent No. column across all sheets.

Assignee lookup priority:
  1. eval_actual_assignee
  2. assignee
  3. Retry with underscores stripped (US_12534530_B2 → US12534530B2)
  4. "Unknown" if all fail

Usage:
    python export_charts.py --drug Axitinib
    python export_charts.py --drug Axitinib --input path/to/custom.json
    python export_charts.py --drug Axitinib --output my_charts.xlsx
    python export_charts.py --drug Axitinib --no_bq   # skip BQ, leave Filed By blank
"""

import argparse
import json
import os
import re
import sys
from pathlib import Path
from typing import Optional

# Load .env if present (python-dotenv)
try:
    from dotenv import load_dotenv
    load_dotenv(dotenv_path=os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"))
except ImportError:
    pass

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

STEP9_OUTPUT_DIR = Path(__file__).parent / "step9_output"
BQ_TABLE         = os.getenv("BQ_PATENT_TABLE",
                             "prj-portfolio-ai-dev.portfolio_data.patent_discovery")


# ─────────────────────────────────────────────────────────────
# BigQuery assignee lookup
# ─────────────────────────────────────────────────────────────

def _strip_underscores(patent_number: str) -> str:
    """US_12534530_B2  →  US12534530B2"""
    return patent_number.replace("_", "")


def fetch_assignees(patent_numbers: list[str]) -> dict[str, str]:
    """
    Query BigQuery for all patent numbers in one batched call.
    Returns {patent_number: assignee_string}.

    Tries each patent number as-is first; if no row found, retries
    with underscores stripped.
    """
    if not patent_numbers:
        return {}

    try:
        from google.cloud import bigquery
    except ImportError:
        print("[BQ] google-cloud-bigquery not installed — Filed By will be blank.")
        return {}

    try:
        sa_path = os.getenv("GCS_SERVICE_ACCOUNT")
        if sa_path:
            from google.oauth2 import service_account
            credentials = service_account.Credentials.from_service_account_file(
                sa_path,
                scopes=["https://www.googleapis.com/auth/bigquery.readonly"],
            )
            client = bigquery.Client(credentials=credentials,
                                     project=credentials.project_id)
            print(f"[BQ] Authenticated via GCS_SERVICE_ACCOUNT: {sa_path}")
        else:
            client = bigquery.Client()
            print("[BQ] GCS_SERVICE_ACCOUNT not set — using application default credentials")
    except Exception as e:
        print(f"[BQ] Could not create BigQuery client: {e} — Filed By will be blank.")
        return {}

    # Build two sets: original and stripped variants
    originals = list(dict.fromkeys(patent_numbers))   # deduplicated, order preserved
    stripped  = {_strip_underscores(p): p for p in originals}  # stripped → original

    # Combined set for a single IN-clause query
    all_variants = list(dict.fromkeys(originals + list(stripped.keys())))
    placeholders = ", ".join(f"'{p}'" for p in all_variants)

    query = f"""
        SELECT patent_number, eval_actual_assignee, assignee
        FROM `{BQ_TABLE}`
        WHERE patent_number IN ({placeholders})
    """

    print(f"[BQ] Querying assignees for {len(originals)} patent(s)...")
    try:
        rows = list(client.query(query).result())
    except Exception as e:
        print(f"[BQ] Query failed: {e} — Filed By will be blank.")
        return {}

    # Index results: patent_number → best assignee value
    bq_index: dict[str, str] = {}
    for row in rows:
        pn    = row.patent_number or ""
        value = row.eval_actual_assignee or row.assignee or ""
        if pn and value:
            bq_index[pn] = value

    print(f"[BQ] Got {len(bq_index)} result(s) from BigQuery.")

    # Map each original patent number to its assignee
    result: dict[str, str] = {}
    for orig in originals:
        if orig in bq_index:
            result[orig] = bq_index[orig]
            continue
        # Try stripped variant
        s = _strip_underscores(orig)
        if s != orig and s in bq_index:
            result[orig] = bq_index[s]
            print(f"[BQ]   {orig} → matched via stripped form '{s}'")
            continue
        # Not found
        print(f"[BQ]   {orig} → no assignee found (Filed By: Unknown)")
        result[orig] = "Unknown"

    return result


# ─────────────────────────────────────────────────────────────
# Excel helpers
# ─────────────────────────────────────────────────────────────

def _border():
    thin = Side(style="thin", color="000000")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _cell(ws, row, col, value, bold=False, wrap=True, font_size=10):
    c = ws.cell(row=row, column=col, value=str(value) if value is not None else "")
    c.font      = Font(name="Arial", bold=bold, size=font_size)
    c.alignment = Alignment(wrap_text=wrap, vertical="top")
    c.border    = _border()
    return c

def _header_row(ws, headers, row, font_size=10):
    for col, h in enumerate(headers, 1):
        _cell(ws, row, col, h, bold=True, font_size=font_size)
    return row + 1

def _col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ─────────────────────────────────────────────────────────────
# Excel builder
# ─────────────────────────────────────────────────────────────

def build_excel(
    drug_name:  str,
    all_charts: list,
    output_path: Path,
    assignees:  dict[str, str],
) -> None:

    def filed_by(patent: str) -> str:
        return assignees.get(patent, "Unknown")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Claim Charts ─────────────────────────────────
    # One row per limitation. All prior art for that limitation
    # (across all grounds) is merged into a single "Prior Art" cell.
    ws1 = wb.create_sheet("Claim Charts")
    _col_widths(ws1, [14, 30, 8, 14, 110])
    r = _header_row(ws1, [
        "Patent No.", "Filed By", "Claim",
        "Limitation ID | Limitation (verbatim)",
        "Prior Art (all references)",
    ], 1)

    for chart_data in all_charts:
        patent = chart_data.get("patent", "")
        fb     = filed_by(patent)

        # Group by claim number, then by limitation_id
        # Collect all prior art passages per (claim, limitation_id) across all grounds
        from collections import defaultdict, OrderedDict

        # Structure: claim_number -> limitation_id -> { ltext, [prior_art_blocks] }
        claim_lim_map: dict = OrderedDict()

        for chart in chart_data.get("charts", []):
            cn        = chart["claim_number"]
            ground_id = chart["ground_id"]
            basis     = chart["basis"]

            if cn not in claim_lim_map:
                claim_lim_map[cn] = OrderedDict()

            for row_data in chart.get("rows", []):
                lid   = row_data.get("limitation_id", "")
                ltext = row_data.get("limitation_text", "")

                if lid not in claim_lim_map[cn]:
                    claim_lim_map[cn][lid] = {
                        "ltext":          ltext,
                        "prior_art_parts": [],
                    }

                # Collect all evidence passages for this limitation in this ground
                for cd in row_data.get("cells", []):
                    ref_id = cd.get("reference_id", "")
                    if ref_id == "—":
                        continue
                    for p in cd.get("passages", []):
                        passage = p.get("passage_verbatim", "")
                        if not passage or "Not disclosed" in passage:
                            continue
                        locus    = p.get("locus", "")
                        url      = p.get("citation_url", "")
                        pub_date = p.get("publication_date", "")
                        score    = p.get("confidence_score", "")
                        rationale = p.get("reads_on_rationale", "")
                        sub_feat  = p.get("sub_feature", "")

                        lines = []
                        header = f"[{ref_id}]"
                        if pub_date:
                            header += f"  ({pub_date})"
                        if score != "":
                            header += f"  Score: {score}"
                        lines.append(header)
                        lines.append(f'"{passage}"')
                        if locus:
                            lines.append(f"Locus: {locus}")
                        if url:
                            lines.append(f"URL: {url}")
                        if rationale:
                            lines.append(f"Reads on: {rationale}")
                        if sub_feat and sub_feat != "full":
                            lines.append(f"Sub-feature: {sub_feat}")

                        block = "\n".join(lines)
                        # Deduplicate — same ref may appear in multiple grounds
                        if block not in claim_lim_map[cn][lid]["prior_art_parts"]:
                            claim_lim_map[cn][lid]["prior_art_parts"].append(block)

        # Now write one row per limitation (one row per unique claim+lim)
        for cn, lims in claim_lim_map.items():
            # Claim header row
            ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
            c = ws1.cell(row=r, column=1,
                         value=f"Patent: {patent}  |  Claim {cn}  |  Filed By: {fb}")
            c.font      = Font(name="Arial", bold=True, size=10)
            c.alignment = Alignment(wrap_text=True, vertical="center")
            c.border    = _border()
            ws1.row_dimensions[r].height = 18
            r += 1

            for lid, data in lims.items():
                ltext = data["ltext"]
                parts = data["prior_art_parts"]

                prior_art_text = (
                    ("\n\n" + "─" * 50 + "\n\n").join(parts)
                    if parts else "Not disclosed — see Gap List"
                )
                lim_cell_text = f"[{lid}]\n{ltext}"

                _cell(ws1, r, 1, patent)
                _cell(ws1, r, 2, fb)
                _cell(ws1, r, 3, cn)
                _cell(ws1, r, 4, lim_cell_text, bold=False)
                _cell(ws1, r, 5, prior_art_text)

                n_lines = prior_art_text.count("\n") + 2
                ws1.row_dimensions[r].height = max(40, min(400, 14 * n_lines))
                r += 1

            r += 1  # blank row between claims/patents

    ws1.freeze_panes = "A2"

    # ── Sheet 2: Coverage Summary ─────────────────────────────
    ws2 = wb.create_sheet("Coverage Summary")
    _col_widths(ws2, [14, 30, 8, 8, 45, 15, 12, 50, 50])
    r = _header_row(ws2, [
        "Patent No.", "Filed By", "Claim", "Ground", "Reference(s)",
        "Covered / Total", "Coverage %", "Uncovered Lim IDs", "Combination Rationale",
    ], 1)

    for chart_data in all_charts:
        patent = chart_data.get("patent", "")
        fb     = filed_by(patent)
        for cs in chart_data.get("coverage_summary", []):
            _cell(ws2, r, 1, patent)
            _cell(ws2, r, 2, fb)
            _cell(ws2, r, 3, cs.get("claim", ""))
            _cell(ws2, r, 4, cs.get("ground", ""))
            _cell(ws2, r, 5, ", ".join(cs.get("references", [])))
            _cell(ws2, r, 6, cs.get("covered_total", ""))
            _cell(ws2, r, 7, f"{cs.get('coverage_pct', 0)}%")
            _cell(ws2, r, 8, ", ".join(cs.get("uncovered_lim_ids", [])) or "—")
            _cell(ws2, r, 9, cs.get("combination_rationale", "—"))
            ws2.row_dimensions[r].height = 30
            r += 1

    ws2.freeze_panes = "A2"

    # ── Sheet 3: Reference List ───────────────────────────────
    ws3 = wb.create_sheet("Reference List")
    _col_widths(ws3, [14, 30, 25, 60, 15, 14, 25, 50])
    r = _header_row(ws3, [
        "Patent No.", "Filed By", "Short Name", "Full Citation",
        "Publication Date", "Pre-Priority?", "Access", "Citation URL",
    ], 1)

    for chart_data in all_charts:
        patent = chart_data.get("patent", "")
        fb     = filed_by(patent)
        for ref in chart_data.get("reference_list", []):
            _cell(ws3, r, 1, patent)
            _cell(ws3, r, 2, fb)
            _cell(ws3, r, 3, ref.get("short_name", ""), bold=True)
            _cell(ws3, r, 4, ref.get("full_citation", ""))
            _cell(ws3, r, 5, ref.get("publication_date", ""))
            _cell(ws3, r, 6, "Yes" if ref.get("pre_priority") else "No")
            _cell(ws3, r, 7, ref.get("access", "publicly accessible"))
            _cell(ws3, r, 8, ref.get("citation_url", ""))
            ws3.row_dimensions[r].height = 25
            r += 1

    ws3.freeze_panes = "A2"

    # ── Sheet 4: Gap List ─────────────────────────────────────
    ws4 = wb.create_sheet("Gap List")
    _col_widths(ws4, [14, 30, 8, 14, 60, 30, 30, 55])
    r = _header_row(ws4, [
        "Patent No.", "Filed By", "Claim", "Limitation ID",
        "Limitation (verbatim)", "Flags", "Recommended Source", "Recommendation",
    ], 1)

    for chart_data in all_charts:
        patent = chart_data.get("patent", "")
        fb     = filed_by(patent)
        for gap in chart_data.get("gap_list", []):
            _cell(ws4, r, 1, patent)
            _cell(ws4, r, 2, fb)
            _cell(ws4, r, 3, gap.get("claim_number", ""))
            _cell(ws4, r, 4, gap.get("limitation_id", ""), bold=True)
            _cell(ws4, r, 5, gap.get("limitation_text", ""))
            raw_flags = gap.get("flags", [])
            flag_strs = [f.get("flag", str(f)) if isinstance(f, dict) else str(f) for f in raw_flags]
            _cell(ws4, r, 6, "; ".join(flag_strs) or "—")
            _cell(ws4, r, 7, gap.get("recommended_source", "—"))
            _cell(ws4, r, 8, gap.get("recommendation", "—"))
            ws4.row_dimensions[r].height = 35
            r += 1

    if r == 2:
        ws4.cell(row=2, column=1, value="No gaps — all limitations covered.")

    ws4.freeze_panes = "A2"

    # ── Sheet 5: Grace Annex ──────────────────────────────────
    ws5 = wb.create_sheet("Grace Annex")
    _col_widths(ws5, [14, 30, 8, 14, 35, 60])
    r = _header_row(ws5, [
        "Patent No.", "Filed By", "Claim", "Limitation ID",
        "Reference (Grace Period)", "Note",
    ], 1)

    for chart_data in all_charts:
        patent = chart_data.get("patent", "")
        fb     = filed_by(patent)
        for ga in chart_data.get("grace_annex", []):
            _cell(ws5, r, 1, patent)
            _cell(ws5, r, 2, fb)
            _cell(ws5, r, 3, ga.get("claim_number", ""))
            _cell(ws5, r, 4, ga.get("limitation_id", ""), bold=True)
            _cell(ws5, r, 5, ga.get("reference_id", ""))
            _cell(ws5, r, 6,
                  "Published within 12 months of priority — "
                  "admissibility case-by-case; counsel to confirm")
            ws5.row_dimensions[r].height = 30
            r += 1

    if r == 2:
        ws5.cell(row=2, column=1, value="No grace-period-only limitations.")

    ws5.freeze_panes = "A2"

    # ── Sheet 6: Source Summary ───────────────────────────────
    # Counts how many prior art references came from each source
    ws6 = wb.create_sheet("Source Summary")
    _col_widths(ws6, [14, 30, 30, 12])

    r = _header_row(ws6, ["Patent No.", "Filed By", "Source", "Count"], 1)

    # Collect source counts per patent + overall
    from collections import Counter
    overall_counts = Counter()

    for chart_data in all_charts:
        patent = chart_data.get("patent", "")
        fb     = filed_by(patent)

        # Count sources from all evidence across all charts for this patent
        patent_counts = Counter()
        for chart in chart_data.get("charts", []):
            for row_data in chart.get("rows", []):
                for cd in row_data.get("cells", []):
                    ref_id = cd.get("reference_id", "")
                    if ref_id == "—":
                        continue
                    for p in cd.get("passages", []):
                        if p.get("passage_verbatim") and "Not disclosed" not in p.get("passage_verbatim", ""):
                            # Infer source from citation_url
                            url = p.get("citation_url", "")
                            if "patents.google" in url or "patent" in ref_id.lower():
                                patent_counts["Google Patents"] += 1
                            elif "pubmed" in url or "PMID" in ref_id:
                                patent_counts["PubMed"] += 1
                            elif "clinicaltrials" in url or "NCT" in ref_id:
                                patent_counts["ClinicalTrials.gov"] += 1
                            elif "medrxiv" in url or "biorxiv" in url:
                                patent_counts["medRxiv / bioRxiv"] += 1
                            else:
                                patent_counts["Other"] += 1

        # Also count from reference_list (more reliable — has explicit source field)
        ref_source_counts = Counter()
        for ref in chart_data.get("reference_list", []):
            source = ref.get("full_citation", "")
            if "Google Patents" in source:
                ref_source_counts["Google Patents"] += 1
            elif "PubMed" in source:
                ref_source_counts["PubMed"] += 1
            elif "ClinicalTrials" in source:
                ref_source_counts["ClinicalTrials.gov"] += 1
            elif "medRxiv" in source or "bioRxiv" in source:
                ref_source_counts["medRxiv / bioRxiv"] += 1
            else:
                ref_source_counts["Other"] += 1

        # Use reference_list counts if available, else passage-inferred counts
        counts = ref_source_counts if ref_source_counts else patent_counts

        for source_name in sorted(counts.keys()):
            _cell(ws6, r, 1, patent)
            _cell(ws6, r, 2, fb)
            _cell(ws6, r, 3, source_name, bold=True)
            _cell(ws6, r, 4, counts[source_name])
            ws6.row_dimensions[r].height = 22
            r += 1

        overall_counts += counts

    # Overall total row
    if overall_counts:
        r += 1  # blank row
        _cell(ws6, r, 1, "", bold=True)
        _cell(ws6, r, 2, "", bold=True)
        ws6.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        _cell(ws6, r, 1, "OVERALL TOTAL", bold=True)
        _cell(ws6, r, 3, "", bold=True)
        _cell(ws6, r, 4, "", bold=True)
        r += 1

        for source_name in sorted(overall_counts.keys()):
            _cell(ws6, r, 1, "")
            _cell(ws6, r, 2, "")
            _cell(ws6, r, 3, source_name, bold=True)
            _cell(ws6, r, 4, overall_counts[source_name])
            ws6.row_dimensions[r].height = 22
            r += 1

        # Grand total
        _cell(ws6, r, 1, "")
        _cell(ws6, r, 2, "")
        _cell(ws6, r, 3, "Total References", bold=True)
        _cell(ws6, r, 4, sum(overall_counts.values()), bold=True)

    ws6.freeze_panes = "A2"

    wb.save(str(output_path))
    print(f"Saved: {output_path}")


# ─────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Export step9 all_final_charts JSON to Excel with Filed By column"
    )
    parser.add_argument("--drug",   "-d", required=True,  help="Drug name")
    parser.add_argument("--input",  "-i", default=None,   help="Path to JSON file (optional)")
    parser.add_argument("--output", "-o", default=None,   help="Output .xlsx path (optional)")
    parser.add_argument("--no_bq",        action="store_true",
                        help="Skip BigQuery lookup; Filed By will show 'Unknown'")
    args = parser.parse_args()

    # Locate input JSON
    if args.input:
        json_path = Path(args.input)
    else:
        safe = re.sub(r"[^a-zA-Z0-9_-]", "_", args.drug)
        json_path = STEP9_OUTPUT_DIR / f"{safe}_all_final_charts.json"

    if not json_path.exists():
        print(f"ERROR: JSON not found: {json_path}", file=sys.stderr)
        print("Run step9.py first, or pass --input with the correct path.")
        sys.exit(1)

    all_charts = json.loads(json_path.read_text(encoding="utf-8"))
    if isinstance(all_charts, dict):
        all_charts = [all_charts]

    print(f"Loaded {len(all_charts)} patent(s) from {json_path.name}")

    # Collect all unique patent numbers across all charts
    patent_numbers = list(dict.fromkeys(
        c.get("patent", "") for c in all_charts if c.get("patent")
    ))
    print(f"Patents: {patent_numbers}")

    # Fetch assignees from BigQuery
    assignees: dict[str, str] = {}
    if not args.no_bq:
        assignees = fetch_assignees(patent_numbers)
    else:
        print("[BQ] Skipped (--no_bq)")

    # Determine output path
    if args.output:
        out_path = Path(args.output)
    else:
        safe = re.sub(r"[^a-zA-Z0-9_-]", "_", args.drug)
        out_path = json_path.parent / f"{safe}_claim_charts.xlsx"

    build_excel(args.drug, all_charts, out_path, assignees)


if __name__ == "__main__":
    main()
